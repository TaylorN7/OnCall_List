from openpyxl import Workbook
import calendar

wb = Workbook()
ws = wb.create_sheet('OnCall Schedule')
ws = wb.active
ws.column_dimensions['A'].width = 35

ws['A1'] = 'Date'
ws['B1'] = 'Engineer'

cns_techs = [
    "Jeff Bertuch",
    "Nick Langley",
    "David Hecker",
    "James Mixon", 
    "Corbin Casper",
    "Tim Rhodes",
    "Taylor Neves",
    "Chad Renfro",
    "Brian Mayorga", 
]

months = [
    "January",
    "February", 
    "March",
    "April",
    "May",
    "June",
    "July",
    "August",
    "September",
    "October",
    "November",
    "December",
]

days = [
    "Monday",
    "Tuesday",
    "Wednesday",
    "Thursday",
    "Friday",
    "Saturday",
    "Sunday",
]

month_val = 1
c = calendar.Calendar()
tech_count = 0
day_count = 0
r = 2
day_list = []

for month in months:
    for i in c.itermonthdates(2020,month_val):
        # print(i)

        if day_count == 7:
            day_count = 0
            tech_count += 1
            r += 1

        if tech_count == 9:
           tech_count = 0
            
        if i in day_list:
            continue

        else:
            if day_count == 0:
                print(i, days[day_count], cns_techs[tech_count], end="\n")
                today = str(i) + " " + str(days[day_count])

                ws.cell(row = r, column = 1).value = today
                ws.cell(row = r, column = 2).value = cns_techs[tech_count]
                r += 1

                tech_count += 1
                day_count += 1
                day_list.append(i)

            elif day_count == 1 or day_count == 2 or day_count == 3:
                print(i, days[day_count], cns_techs[tech_count], end="\n")
                today = str(i) + " " + str(days[day_count])

                ws.cell(row = r, column = 1).value = today
                ws.cell(row = r, column = 2).value = cns_techs[tech_count]
                r += 1

                tech_count += 1
                day_count += 1
                day_list.append(i)

            elif day_count == 4 or day_count == 5:
                print(i, days[day_count], cns_techs[tech_count], end="\n")
                today = str(i) + " " + str(days[day_count])

                ws.cell(row = r, column = 1).value = today
                ws.cell(row = r, column = 2).value = cns_techs[tech_count]
                r += 1
                day_count += 1
                day_list.append(i)

            elif day_count == 6:
                print(i, days[day_count], cns_techs[tech_count], end="\n")
                today = str(i) + " " + str(days[day_count])

                ws.cell(row = r, column = 1).value = today
                ws.cell(row = r, column = 2).value = cns_techs[tech_count]
                r += 1
                day_count += 1
                day_list.append(i)

    
    month_val += 1


wb.save('2020_OnCall_W_JB.xlsx')